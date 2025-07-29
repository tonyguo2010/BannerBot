import os

from openpyxl import load_workbook
from openpyxl import Workbook
import ScriptGenerator
from datetime import datetime

employees = []

def load_list_from_sheet_by_col(filename : str, sheet_index : int, col_index: int):
    result = []

    workbook = load_workbook(filename)

    worksheet = workbook[workbook.sheetnames[sheet_index]]
    # header = worksheet.cell(1, col_index).value
    for row in range(1, worksheet.max_row + 1):
        result.append(worksheet.cell(row, col_index).value)

    return result

def init_headers(filename : str):
    headers = []
    workbook = load_workbook(filename)

    worksheet = workbook[workbook.sheetnames[0]]

    for i, row in enumerate(worksheet.iter_rows(), start=1):
        if i == 1:
            continue

        for cell in row:
            header = worksheet.cell(1, cell.column).value
            headers.append(header)

        break

    return headers

def generate_side_script(filename : str):
    scripts = []

    workbook = load_workbook(filename)

    worksheet = workbook[workbook.sheetnames[0]]

    for i, row in enumerate(worksheet.iter_rows(), start=1):
        if i == 1:
            continue
        detail = {}

        for cell in row:
            header = worksheet.cell(1, cell.column).value
            detail[header] = cell.value

        employees.append(detail)

    for employee in employees:
        script_name = ScriptGenerator.convert_row_to_script(employee)
        scripts.append(script_name)

    return scripts

def save_results():
    wb = Workbook()
    ws = wb.active
    ws.title = "PDOC results"

    # Add data to cells
    ws['A1'] = "Employee ID"
    # ws['B1'] = "Name"
    ws['B1'] = "Federal tax deduction on bonus"
    ws['C1'] = "Provincial tax deduction on bonus"
    ws['D1'] = "Total tax deduction on bonus"

    rows = []
    for employee in employees:
        try:
            items = []
            items.append(employee['SPRIDEN_ID'])
            # items.append(employee['EMPL_NAME'])
            items.append(employee['FED_BONUS_TAX'])
            items.append(employee['ONT_BONUS_TAX'])
            items.append(employee['TOT_BONUS_TAX'])
            rows.append(items)
        except:
            print('skipped ' + employee['SPRIDEN_ID'])

    for row in rows:
        ws.append(row)
        # change the remove to here
        # print(row)
        os.remove(row[0]+'.side')

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"output_{timestamp}.xlsx"
    wb.save(filename)